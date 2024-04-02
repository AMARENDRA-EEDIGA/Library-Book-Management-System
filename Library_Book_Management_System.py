
import tkinter as tk
from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox,Toplevel
from PIL import Image, ImageTk
import openpyxl
import pathlib
import os

    
    
def validate():        
    if college_email_entry.get() == "info" and Librarian_ID_entry.get() == "svdc":   
    # if college_email_entry.get() == "info@svdegreecollege.ac.in" and Librarian_ID_entry.get() == "svdc3012":
        messagebox.showinfo("Success","Successfully logged in")
        # custom_message_box = messagebox.showinfo("Login successful", "Successfully logged in.")
        # custom_message_box.geometry("500x300")
        
        # Hide the login window
        new_window.withdraw()
        
        # Create a new window with the same geometry settings
        root = tk.Toplevel()
        root.geometry("1000x600+50+50")
    
        background = "#06283D"
        skyblue = "#68ddfa"
        
        root.title("Library Book Management System")
        root.config(bg=background)
        
        file_path = 'LibraryBooks.xlsx'
        
        if not pathlib.Path(file_path).exists():
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            titles = ["Serial Number", "Title", "Author", "Year", "Publisher", "Price", "Quantity", "Date"]
            
            col = 1
            for title in titles:
                sheet.cell(row=1, column=col, value=title)
                col += 1
            
            workbook.save(file_path)
        
        
        def Exit():
            root.destroy()
        
        def showimage():
            global img
            global photo2
            # global lbl
            filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select image file",
                                                  filetypes=(("JPG File", "*.jpg"), ("PNG File", "*.png"), ("All files", "*.*")))
            if filename:
                img = Image.open(filename)
                resize_img = img.resize((200, 200))  
                photo2 = ImageTk.PhotoImage(resize_img)
                lbl.config(image=photo2)
                lbl.image = photo2  
        
        def clear():
            global img
        
            Title.set('')
            Author.set('')
            Year.set('')
            Publisher.set('')
            Price.set('')
            Quantity.set('')  
            SerialNumber.set('')

            img = Image.open("Images/Upload photo.png")
            resize_img = img.resize((200, 200))
            photo2 = ImageTk.PhotoImage(resize_img)
            # lbl.config(image=photo2)
            lbl.config(image=photo2)
            lbl.image = photo2  
            # saveButton.config(state="normal")  
        
        def Save():
            T1 = Title.get()
            A1 = Author.get()
            Y1 = Year.get()
            Pub1 = Publisher.get()
            Pri1 = Price.get()
            Q1 = Quantity.get()
            S1 = SerialNumber.get()
            D1 = Date.get()
        
            if T1=="" or A1=="" or Y1=="" or Pub1=="" or Pri1=="" or Q1=="" or S1=="":
                messagebox.showerror("Error:", "Few Data is Missing.")
            else:
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                sheet.append([S1, T1, A1, Y1, Pub1, Pri1, Q1, D1])
                wb.save(file_path)
                img.save(os.path.join("Book Images", f"{S1}.jpg"))
                messagebox.showinfo("Success:", "Data saved successfully.")
                clear()
        
        def search():
            text = Search.get().strip()  
            clear()  

            try:
                row_num = int(text) 
            except ValueError:
                messagebox.showerror("Invalid", "Invalid Serial Number.")
                return

            file = openpyxl.load_workbook(file_path)
            sheet = file.active

            found = False
            for row in sheet.iter_rows(values_only=True):
                if row[0] == row_num:  
                    found = True
            
                    SerialNumber.set(row[0])
                    Title.set(row[1])
                    Author.set(row[2])
                    Year.set(row[3])
                    Publisher.set(row[4])
                    Price.set(row[5])
                    Quantity.set(row[6])
                    Date.set(row[7])
                
                    img_path = os.path.join("Book Images/", str(row[0]) + ".jpg")
                    if os.path.exists(img_path):
                        img = Image.open(img_path)
                        reSized_image = img.resize((200, 200))
                        photo2 = ImageTk.PhotoImage(reSized_image)
                        lbl.config(image=photo2)
                        lbl.image = photo2
                    else:
                        messagebox.showerror("Image Error", "Image not found.")
                    break  
        
            if not found:
                messagebox.showerror("Login failed:","Invalid Email or ID")
        
        # fg = foreground color of the text(text-color)
        Label(root, text="LIBRARY BOOK MANAGEMENT", height=2, bg="green", fg="#fff", font='arial 20 bold').pack(side=TOP,fill=X)
        
        Search = StringVar()
        Entry(root, textvariable=Search, width=15, font="arial 13").place(x=760, y=21)
        
        srch = Button(root, text="Search", compound=LEFT, width=10, height=1, bg=skyblue, font="arial 8 bold", command=search)
        srch.place(x=900, y=21)
        
        Label(root, text="Serial-Number:", font="arial 13", fg="orange", bg=background).place(x=30, y=80)
        Label(root, text="Date:", font="arial 13", fg="orange", bg=background).place(x=550, y=80)
        Button(root, text="Clear Data", width=18, height=1, font="arial 12 bold", bg="#f0687c", command=clear).place(x=760, y=80)
        
        SerialNumber = IntVar()
        SerNum_entry = Entry(root, textvariable=SerialNumber, width=15, font="arial 10")
        SerNum_entry.place(x=150, y=80)
        
        Date = StringVar()
        today = date.today()
        d1 = today.strftime("%d/%m/%Y")
        date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
        date_entry.place(x=600, y=80)
        Date.set(d1)
        
        obj = LabelFrame(root, text="Book's Details", font=20, width=680, height=400, bg="gray", fg="yellow", relief=GROOVE)
        obj.place(x=30, y=150)
        
        Label(obj, text="Title:", font="arial 13", bg="gray", fg="black").place(x=40, y=50)
        Label(obj, text="Author:", font="arial 13", bg="gray", fg="black").place(x=40, y=100)
        Label(obj, text="Year:", font="arial 13", bg="gray", fg="black").place(x=40, y=150)
        Label(obj, text="Publisher:", font="arial 13", bg="gray", fg="black").place(x=40, y=200)
        Label(obj, text="Price:", font="arial 13", bg="gray", fg="black").place(x=40, y=250)
        Label(obj, text="Quantity:", font="arial 13", bg="gray", fg="black").place(x=40, y=300)
        
        Title = StringVar()
        Title_entry = Entry(obj, textvariable=Title, width=30, font="arial 10")
        Title_entry.place(x=150, y=50)
        
        Author = StringVar()
        Author_entry = Entry(obj, textvariable=Author, width=30, font="arial 10")
        Author_entry.place(x=150, y=100)
        
        Year = StringVar()
        Year_entry = Entry(obj, textvariable=Year, width=30, font="arial 10")
        Year_entry.place(x=150, y=150)
        
        Publisher = StringVar()
        Publisher_entry = Entry(obj, textvariable=Publisher, width=30, font="arial 10")
        Publisher_entry.place(x=150, y=200)
        
        Price = StringVar()
        Price_entry = Entry(obj, textvariable=Price, width=30, font="arial 10")
        Price_entry.place(x=150, y=250)
        
        Quantity = StringVar()
        Quantity_entry = Entry(obj, textvariable=Quantity, width=30, font="arial 10")
        Quantity_entry.place(x=150, y=300)
        
        f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
        f.place(x=750, y=150)
        
        img = Image.open("Images/Upload photo.png")
        resize_img = img.resize((200, 200))
        photo2 = ImageTk.PhotoImage(resize_img)
        lbl = Label(f, image=photo2)
        lbl.place(x=0, y=0)
        
        Button(root, text="Upload", width=18, height=1, font="arial 12 bold", bg="lightblue", command=showimage).place(x=760, y=370)
        
        saveButton = Button(root, text="Save", width=18, height=1, font="arial 12 bold", bg="lightgreen", command=Save)
        saveButton.place(x=760, y=440)
        
        Button(root,text="Exit",width=18,height=1,font="arial 12 bold",bg="red",command=Exit).place(x=760,y=510)
        
        
        root.mainloop()
        

        # Show the new window
        root.deiconify()
    else:
        messagebox.showerror("Login failed")



new_window = tk.Tk()

new_window.title("Login Page")
new_window.geometry('800x600+200+100')

bg_image = Image.open("Images/geometric-background.png")
bg_image = bg_image.resize((new_window.winfo_screenwidth(), new_window.winfo_screenheight()))
bg_image = ImageTk.PhotoImage(bg_image)

background_label = tk.Label(new_window, image=bg_image)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

f = Label(new_window, bd=3,image=bg_image, width=300, height=330, relief=tk.GROOVE, highlightthickness=0)
f.place(x=250, y=140)

username_label = tk.Label(f, text="College Email",width=20,font="arial 13", bg="gray", fg="white")
username_label.place(x=60,y=40)

college_email_entry = tk.Entry(f,width=20,font="arial 12 bold")
college_email_entry.place(x=60,y=80)

password_label = tk.Label(f, text="Librarian ID",width=20,font="arial 13", bg="gray", fg="white")
password_label.place(x=60,y=140)

Librarian_ID_entry = tk.Entry(f, show="*",width=20,font="arial 12 bold")
Librarian_ID_entry.place(x=60,y=180)

login_button = tk.Button(f, text="Login",width=18,height=1,font="arial 12 bold",bg="#68ddfa",bd=3, command=validate)
login_button.place(x=60,y=230)

new_window.mainloop()

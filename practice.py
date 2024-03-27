from tkinter import *
from tkinter import messagebox
import openpyxl

root = Tk()
root.title("Practical Session")
root.geometry("1000x300+50+50")
root.config(bg="#68ddfa")

file_path = "C:\\Users\\Admin\\OneDrive\\Desktop\\GUI\\Excel_file.xlsx"

wb = openpyxl.load_workbook(file_path)
sheet = wb.active
sheet.cell(row=1,column=1,value="Name")
sheet.cell(row=1,column=2,value="Age")
wb.save(file_path)


def Save():
    N1 = Name.get()
    A1 = Age.get()
    
    if N1 == "" or A1 == "":
        messagebox.showerror("Error:", "Few Data is Missing.")
    else:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        sheet.append([N1, A1])
        wb.save(file_path)
        messagebox.showinfo("Success:", "Data saved successfully.")
        
        
def search():
    text = Search.get()

    file = openpyxl.load_workbook(file_path)
    sheet = file.active

    found = False
    for row in sheet.iter_rows(values_only=True):
        if row[0] == text:
            Name.set(row[0])
            Age.set(row[1])
            found = True
            break

    if not found:
        messagebox.showinfo("Search", "No matching record found.")


def clear():
    Name.set("")
    Age.set("")
    

Search = StringVar()
search_Entry = Entry(root, textvariable=Search, width=15, font="arial 13")
search_Entry.place(x=760, y=21)
search_Entry.insert(0,"Search Name")

srch = Button(root, text="Search", compound=LEFT, width=10, height=1, font="arial 8 bold", command=search)
srch.place(x=900, y=21)

obj = LabelFrame(root, font=20, width=750, height=200, bg="gray", fg="yellow", relief=GROOVE)
obj.place(x=100, y=50)

Label(obj, text="Name:", font="arial 13", bg="gray", fg="black").place(x=40, y=50)
Label(obj, text="Age:", font="arial 13", bg="gray", fg="black").place(x=40, y=100)

Name = StringVar()
Name_entry = Entry(obj, textvariable=Name, width=30, font="arial 10")
Name_entry.place(x=150, y=50)

Age = StringVar()
Age_entry = Entry(obj, textvariable=Age, width=30, font="arial 10")
Age_entry.place(x=150, y=100)

Button(root, text="Reset", width=15, height=1, font="arial 12 bold", bg="red", command=clear).place(x=660, y=200)
saveButton = Button(root, text="Save", width=15, height=1, font="arial 12 bold", bg="lightgreen", command=Save)
saveButton.place(x=660, y=100)

root.mainloop()

# from tkinter import *

# def on_entry_click(event):
#     if search_entry.get() == "Enter text to search...":
#         search_entry.delete(0, END)
#         search_entry.config(fg='black')

# def on_focus_out(event):
#     if search_entry.get() == "":
#         search_entry.insert(0, "Enter text to search...")
#         search_entry.config(fg='grey')

# root = Tk()
# root.geometry("300x100")

# search_entry = Entry(root, width=30, fg='grey')
# search_entry.insert(0, "Enter text to search...")
# search_entry.bind('<FocusIn>', on_entry_click)
# search_entry.bind('<FocusOut>', on_focus_out)
# search_entry.pack(pady=10)

# root.mainloop()
